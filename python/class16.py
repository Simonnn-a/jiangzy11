# _*_ conding:utf-8 _*-
# @Time:2020/12/1 &{TIME}
'''
    excel的数据驱动实现：
        1. excel操作流程：先操作workbook，再操作sheet，再操作cell（value）
        2. 对于模块的api建议在使用前可以百度了解一下
        3. excel操作结束后记得close
'''
# 导入模块
import openpyxl

# excel操作流程
# 操作工作簿：指定文件路径，进行文件读取，类似于open()函数的文件读取操作
excel = openpyxl.load_workbook('../data/xxx.xlsx')
print(excel)
# 获取sheet：基于工作簿来获取sheet页
names = excel.sheetnames
print(names)
print(type(names))
for name in names:
    print(name)
# 操作单元格cell：指定sheet页，再进行操作
sheet = excel['Sheet3']
print(sheet)
# 获取单元格内容
# print(sheet.values)
# for value in sheet.values:
#     print(value[0])
# 指定单元格进行内容获取
# value = sheet['A5'].value
# print(value)
# 获取行：最大行数
rows = sheet.max_row
print(rows)
# 获取列：最大列数
column = sheet.max_column
print(column)
# 写入前一定确保被操作文件处于关闭状态，否则会报错
# 写入：基于单元格来进行写入
sheet['H1'] = 'asd'
# 写入之后一定记得保存：保存对象是excel，而不是sheet或者cell
excel.save('../data/xxx.xlsx')
# 在所有操作结束之后，记得释放
excel.close()
'''
    基于excel数据驱动来实现自动化测试
    既然关键字已封装，我可否通过对数据的读取，让它自行执行对应的关键对象？
'''
import openpyxl

# 访问excel内容
from class07.key_word_web.keyword_web import WebKeys

excel = openpyxl.load_workbook('../data/xxx.xlsx')
sheet = excel['Sheet3']
# 读取excel内容，实现文件驱动自动化执行
for value in sheet.values:
    # 定义一个字典参数，用于接收excel中的所有参数内容
    args = {}
    # 定位方法
    args['name'] = value[2]
    # 定位路径
    args['value'] = value[3]
    # 输入内容
    args['txt'] = value[4]
    # 预期结果
    args['expect'] = value[6]
    # 基于A列进行判断是否为测试用例
    if type(value[0]) is int:
        '''
            在读取关键字时，分为几类情况：
                1. 关键字驱动类的实例化
                2. 断言类型的关键字
        '''
        # 判断是否实例化
        if value[1] == 'open_browser':
            wk = WebKeys(value[4])
        # 非特殊关键字，通过反射机制实现
        else:
            getattr(wk, value[1])(**args)
            # getattr(wk, 'open')(**args)
            # wk.open(**args)
            # 原有的open函数是 def open(self,url):
            '''
                如果是open，只需要传入value[4]
                如果是input，则需要传入value[2],value[3],value[4]
                如果是click，则需要传入value[2],value[3]
                **args不定长传参，不定义参数数量的多少，说白了就是一个字典格式
                但是字典是键值对形式存在
            '''

