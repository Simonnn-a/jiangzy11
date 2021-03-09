# _*_ conding:utf-8 _*-
# @Time:2020/12/4 &{TIME}
import openpyxl
from openpyxl.styles import PatternFill, Font

from key_excel import WebKeys
from python.test_log import DemoLog
log=DemoLog().txt_log()
excel = openpyxl.load_workbook('ecl/shopping.xlsx')
sheet=excel['购物车']
log.info('获取{0}内容成功，现在开始执行自动化测试......'.format(sheet))
for value in sheet.values:
    args={}
    args['name']=value[2]
    args['value']=value[3]
    args['txt']=value[4]
    args['expect']=value[6]
    if type(value[0]) is int:
        if value[1]=='browser':
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            wd = WebKeys(value[4])
        elif 'assert' in value[1]:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            status=getattr(wd, value[1])(**args)
            if status:
                sheet.cell(row=value[0]+1, column=8).value='Pass'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='AACF91')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True) #FONT导入  oppyexcel.stylee.....
            else:
                sheet.cell(row=value[0] + 1, column=8).value = 'Failed'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
            excel.save('ecl/shopping.xlsx')
        else:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            getattr( wd, value[1])(**args)

excel.close()