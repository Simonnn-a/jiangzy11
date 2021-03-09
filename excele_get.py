#导入模块
import openpyxl

"""
读写操作：
1.openpvxl.load_workbook 打开一个excel
2.sheetnames获取所有的sheet名字
3.values获取当前sheet的所有单元格内容
弊端：绝大部分函数都需要手动编写，没有自动补全。
"""

#读取excel内容操作
#基于excel文件路径，打开excel,默认打开的文件都是workbook
excel = openpyxl.load_workbook('webuikeys.xlsx')
print(excel)
# #获取excel的sheet
names = excel.sheetnames
for name in names:
    print(name)
sheet = excel[names[0]]
print(sheet)
#获取sheet的内容
for value in sheet.values:
    print(value)

#excel写入
sheet['A3'] = 'sssss'
sheet['G10'] = 'asdfwedczx'
#保存机制，如果有写入，必须先关必文件，再执行操作
excel.save('webuikeys.xlsx')
#释放资源
excel.close()









